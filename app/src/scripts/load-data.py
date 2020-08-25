import openpyxl
from datetime import datetime
import json

file = "2020 תוכנית הדרכות בטיחות.xlsx"
file2 = "Production training list.xlsx"

wb = openpyxl.load_workbook(file)
wb2 = openpyxl.load_workbook(file2)

ws = wb["רשימת עובדים"]
ws2 = wb2["List of workers"]
ws3 = wb2["Departments Vs Trainings"]
list_of_departments_sheet = wb2["List of Departments"]


workers = []

worker = {}

# this is the correct order of the fields in the excel file

worker_fields = {
    "workerId": "",
    "lastName": "",
    "firstName": "",
    "birthDate": "",
    "age": "",
    "sex": "",
    "startDate": "",
    "idNumber": "",
    "phoneNumber": "",
    "phoneNumber2": "",
    "city": "",
    "address": "",
    "zipCode": "",
    "departmentHE": "",
    "departmentEN": "",
    "roleEmergency": "",
    "role": "",
}


def date_format(date_string):
    return date_string.strftime("%Y-%m-%d") if type(date_string) is datetime else "N/A"


def set_worker_dates(worker):
    worker["birthDate"] = date_format(worker["birthDate"])
    worker["startDate"] = date_format(worker["startDate"])


def set_worker_team(worker):
    worker["team"] = "N/A"
    for row in ws2.iter_rows(min_row=2):
        if row[0].value is not None:
            row_num = row[0].row
            worker_id = ws2[f"A{row_num}"].value
            worker_team = ws2[f"D{row_num}"].value
            if worker["workerId"] == worker_id:
                worker["team"] = worker_team
                print(f"worker: {worker_id} team: {worker['team']}")
                return


def set_worker_trainings(worker, row):
    row_num = row[0].row
    trainings = [
        # IOS
        {
            "name": ws["R1"].value,
            "isRequired": ws[f"R{row_num}"].value,
            "status": ws[f"S{row_num}"].value,
            "completionDate": date_format(ws[f"T{row_num}"].value),
            "expiryDate": "N/A",
            "trainerName": "N/A"
        },
        # IOS Capstone
        {
            "name": ws["U1"].value,
            "isRequired": ws[f"U{row_num}"].value,
            "status": ws[f"V{row_num}"].value,
            "completionDate": date_format(ws[f"W{row_num}"].value),
            "expiryDate": "N/A",
            "trainerName": "N/A"
        },
        # Ergonomic training
        {
            "name": "Ergonomic",
            "isRequired": ws[f"X{row_num}"].value,
            "status": ws[f"Y{row_num}"].value,
            "completionDate": date_format(ws[f"Z{row_num}"].value),
            "expiryDate": "N/A",
            "trainerName": "N/A"
        },
        # GEMP training
        {
            "name": "GEMP",
            "isRequired": ws[f"AA{row_num}"].value,
            "status": ws[f"AB{row_num}"].value,
            "completionDate": date_format(ws[f"AC{row_num}"].value),
            "expiryDate": "N/A",
            "trainerName": "N/A"
        },

        # רשיון מלגזנים
        {
            "name": "רשיון מלגזנים",
            "isRequired": ws[f"AD{row_num}"].value,
            "status": ws[f"AE{row_num}"].value,
            "completionDate": "N/A",
            "expiryDate": date_format(ws[f"AF{row_num}"].value),
            "trainerName": "N/A"
        },
        # רענון מלגזנים
        {
            "name": "רענון מלגזנים",
            "isRequired": ws[f"AG{row_num}"].value,
            "status": ws[f"AH{row_num}"].value,
            "completionDate": "N/A",
            "expiryDate": date_format(ws[f"AI{row_num}"].value),
            "trainerName": "N/A"
        },
        # מלגזת יד אדם הולך
        {
            "name": "רענון מלגזת יד אדם הולך",
            "isRequired": ws[f"AJ{row_num}"].value,
            "status": ws[f"AK{row_num}"].value,
            "completionDate": date_format(ws[f"AL{row_num}"].value),
            "expiryDate": date_format(ws[f"AN{row_num}"].value),
            "trainerName": "N/A"
        },
        # עבודה בגובה - מבוא
        {
            "name": "עבודה בגובה מבוא",
            "isRequired": ws[f"AO{row_num}"].value,
            "status": ws[f"AP{row_num}"].value,
            "completionDate": date_format(ws[f"AQ{row_num}"].value),
            "expiryDate": "N/A",
            "trainerName": "N/A"
        },
        # עבודה בגובה סולמות גג סל הרמה
        {
            "name": "עבודה בגובה סולמות גג סל הרמה",
            "isRequired": ws[f"AR{row_num}"].value,
            "status": ws[f"AS{row_num}"].value,
            "completionDate": date_format(ws[f"AT{row_num}"].value),
            "expiryDate": "N/A",
            "trainerName": "N/A"
        },
        # הדרכת ניידת בטיחות
        {
            "name": "ניידת בטיחות",
            "isRequired": ws[f"AU{row_num}"].value,
            "status": ws[f"AV{row_num}"].value,
            "completionDate": date_format(ws[f"AW{row_num}"].value),
            "expiryDate": "N/A",
            "trainerName": "N/A"
        },
        # צוות חירום אש + תרגיל
        {
            "name": "צוות חירום אש + תרגיל",
            "isRequired": ws[f"AY{row_num}"].value,
            "status": ws[f"AZ{row_num}"].value,
            "completionDate": date_format(ws[f"BA{row_num}"].value),
            "expiryDate": "N/A",
            "trainerName": "צחי סגל זמן תגובה"
        },
        # החייאה ועזרה ראשונה
        {
            "name": "החייאה ועזרה ראשונה",
            "isRequired": ws[f"BB{row_num}"].value,
            "status": ws[f"BC{row_num}"].value,
            "completionDate": date_format(ws[f"BD{row_num}"].value),
            "expiryDate": "N/A",
            "trainerName": ws[f"BE{row_num}"].value
        }
    ]
    worker["trainings"] = trainings


def set_all_workers():
    for row in ws.iter_rows(min_row=2):
        worker = {}
        if row[0].value is not None:
            for i, key in enumerate(worker_fields.keys()):
                worker[key] = row[i].value
            set_worker_dates(worker)
            set_worker_trainings(worker, row)
            worker["team"] = "N/A"
            workers.append(worker)


def get_clean_string(string):
    return " ".join(string.split())


def get_worker(worker_id):
    for worker in workers:
        if worker["workerId"] == worker_id:
            return worker
    return None


def set_production_workers():
    departments_trainings = get_departments_vs_trainings()
    for row in ws2.iter_rows(min_row=2):
        if row[0].value is not None:
            row_num = row[0].row
            worker_id = ws2[f"A{row_num}"].value
            worker = get_worker(worker_id)
            # if not worker:
            #     worker = worker_fields
            #     worker["workerId"] = ws2[f"A{row_num}"].value
            #     worker["firstName"] = ws2[f"B{row_num}"].value
            #     worker["lastName"] = ws2[f"C{row_num}"].value
            #     workers.append(worker)
            if worker:
                worker["team"] = get_clean_string(ws2[f"D{row_num}"].value)
                set_production_worker_trainings(worker, departments_trainings)

def set_production_worker_trainings(worker, departments_trainings):
    # teams_trainings = get_departments_vs_trainings()
    worker_department = worker["team"]
    if worker_department in departments_trainings.keys():
        print(worker_department)
        department_required_trainings = departments_trainings[f"{worker_department}"]
        for training in department_required_trainings:
            worker["trainings"].append(
                {
                    "name": training,
                    "isRequired": "נדרש",
                    "status": "השתתף",
                    "completionDate": "2020-01-01",
                    "expiryDate": "2021-01-01",
                    "trainerName": "N/A"
                })

def get_departments_vs_trainings():
    departments_trainings = {}
    for row in ws3.iter_rows(min_row=2, max_row=18):
        row_num = row[0].row
        dept_name = get_clean_string(ws3[f"A{row_num}"].value)
        trainings = []
        for cell in row[1:]:
            column_letter = cell.column_letter
            training_required = cell.value # 1 or 0
            if(training_required == 1):
                training_name = get_clean_string(ws3[f"{column_letter}1"].value)
                trainings.append(training_name)
        departments_trainings[dept_name] = trainings
    return departments_trainings
def main():
    set_all_workers()
    set_production_workers()
    with open("workers_data.json", "w", encoding='utf-8') as f:
        json.dump(workers, f, ensure_ascii=False, indent=4)


main()
